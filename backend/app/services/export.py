"""Export of the timeline and project registry to XLSX/CSV — a familiar spreadsheet format."""

import csv
import io
import uuid
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Project
from app.domain.calendar import date_range, is_weekend
from app.domain.capacity import load_calendar_context
from app.domain.categories import CATEGORY_WEIGHTS, XLSX_LETTER

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# formula injection guard: a value starting with =,+,-,@ or a control character
# is treated as a formula by Excel/LibreOffice. Escape with a leading apostrophe.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def _safe(value):
    if isinstance(value, str) and value.startswith(_FORMULA_LEAD):
        return "'" + value
    return value

HEADER_FILL = PatternFill("solid", fgColor="1D2023")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WEEKEND_FILL = PatternFill("solid", fgColor="EDEDED")
OVERLOAD_FILL = PatternFill("solid", fgColor="F5C6C6")


async def _timeline_rows(
    db: AsyncSession, workspace_id: uuid.UUID, date_from: date, date_to: date
):
    members, allocations, absences, non_working = await load_calendar_context(
        db, workspace_id, date_from, date_to
    )
    nw_days = {n.day for n in non_working}
    projects = {
        p.id: p
        for p in (
            await db.execute(
                select(Project).where(Project.workspace_id == workspace_id)
            )
        )
        .scalars()
        .all()
    }
    # weekends are not exported — same as on the timeline; holidays stay
    days = [d for d in date_range(date_from, date_to) if not is_weekend(d)]
    alloc_map: dict[tuple, list] = {}
    for a in allocations:
        alloc_map.setdefault((a.member_id, a.day), []).append(a)

    # overload: the member's category weights for a day sum to more than their capacity
    day_sums: dict[tuple[uuid.UUID, date], Decimal] = {}
    for a in allocations:
        key = (a.member_id, a.day)
        day_sums[key] = day_sums.get(key, Decimal(0)) + CATEGORY_WEIGHTS[a.category]
    capacity = {m.id: Decimal(str(m.capacity_per_day)) for m in members}
    overload_days = {
        key for key, s in day_sums.items() if s > capacity.get(key[0], Decimal(1))
    }

    rows = []
    for m in members:
        member_projects: dict[uuid.UUID, dict[date, str]] = {}
        for a in allocations:
            if a.member_id != m.id:
                continue
            member_projects.setdefault(a.project_id, {})[a.day] = a.category
        for p_id, day_map in sorted(
            member_projects.items(), key=lambda kv: projects[kv[0]].code if kv[0] in projects else ""
        ):
            code = projects[p_id].code if p_id in projects else "?"
            rows.append(
                {
                    "member": m.name,
                    "member_id": m.id,
                    "project": code,
                    "cells": [day_map.get(d) for d in days],
                }
            )
        if not member_projects:
            rows.append(
                {
                    "member": m.name,
                    "member_id": m.id,
                    "project": "—",
                    "cells": [None] * len(days),
                }
            )
    return rows, days, nw_days, overload_days


async def export_timeline_xlsx(
    db: AsyncSession, workspace_id: uuid.UUID, date_from: date, date_to: date
) -> bytes:
    rows, days, nw_days, overload_days = await _timeline_rows(
        db, workspace_id, date_from, date_to
    )
    wb = Workbook()
    sh = wb.active
    sh.title = "Timeline"
    sh.cell(1, 1, "Team member").font = HEADER_FONT
    sh.cell(1, 2, "Project").font = HEADER_FONT
    sh.cell(1, 1).fill = HEADER_FILL
    sh.cell(1, 2).fill = HEADER_FILL
    for i, d in enumerate(days, start=3):
        cell = sh.cell(1, i, f"{DAY_NAMES[d.weekday()]} {d.strftime('%d.%m')}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        sh.column_dimensions[get_column_letter(i)].width = 9
    sh.column_dimensions["A"].width = 24
    sh.column_dimensions["B"].width = 14

    for r, row in enumerate(rows, start=2):
        sh.cell(r, 1, _safe(row["member"]))
        sh.cell(r, 2, _safe(row["project"]))
        for c, (d, value) in enumerate(zip(days, row["cells"]), start=3):
            cell = sh.cell(r, c, XLSX_LETTER[value] if value else None)
            if is_weekend(d) or d in nw_days:
                cell.fill = WEEKEND_FILL
            elif value and (row["member_id"], d) in overload_days:
                cell.fill = OVERLOAD_FILL
            cell.alignment = Alignment(horizontal="center")

    legend = (
        "B — background (0.25) · H — half day (0.5) · "
        "M — most of the day (0.75) · F — full day (1) · red — day overload"
    )
    sh.cell(len(rows) + 3, 1, legend)

    sh.freeze_panes = "C2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_timeline_csv(
    db: AsyncSession, workspace_id: uuid.UUID, date_from: date, date_to: date
) -> str:
    rows, days, _, _ = await _timeline_rows(db, workspace_id, date_from, date_to)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ["Team member", "Project"] + [d.isoformat() for d in days]
    )
    for row in rows:
        writer.writerow(
            [_safe(row["member"]), _safe(row["project"])]
            + [("" if v is None else v) for v in row["cells"]]
        )
    return buf.getvalue()


HEALTH_LABEL = {"green": "🟢", "amber": "🟡", "red": "🔴"}
LIFECYCLE_LABEL = {
    "active": "Active",
    "support": "Support",
    "paused": "Paused",
    "finished": "Finished",
}


async def export_projects_xlsx(db: AsyncSession, workspace_id: uuid.UUID) -> bytes:
    projects = (
        (
            await db.execute(
                select(Project)
                .where(Project.workspace_id == workspace_id, Project.deleted_at.is_(None))
                .order_by(Project.code)
            )
        )
        .scalars()
        .all()
    )
    wb = Workbook()
    sh = wb.active
    sh.title = "Projects"
    headers = ["Code", "Name", "Status", "Health", "Weekly update"]
    widths = [10, 36, 14, 10, 56]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = sh.cell(1, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        sh.column_dimensions[get_column_letter(i)].width = w
    for r, p in enumerate(projects, start=2):
        sh.cell(r, 1, _safe(p.code))
        sh.cell(r, 2, _safe(p.name))
        sh.cell(r, 3, LIFECYCLE_LABEL.get(p.lifecycle, p.lifecycle))
        sh.cell(r, 4, HEALTH_LABEL.get(p.health, p.health))
        sh.cell(r, 5, _safe(p.weekly_update))
    sh.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
